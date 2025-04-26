from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.db.models import Count
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin

import xlwt
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors

from .models import ElectionDistrict, CommissionMember, Representative, Observer
from .forms import ElectionDistrictForm, CommissionMemberForm, RepresentativeForm, ObserverForm
from .filters import ElectionDistrictFilter, CommissionMemberFilter, RepresentativeFilter, ObserverFilter



from .models import CommissionMember, DISTRICT_CHOICES
class CommissionMemberListView(LoginRequiredMixin,ListView):
    model = CommissionMember
    template_name = 'elections/commission_members/list.html'
    context_object_name = 'members'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()
        self.filterset = CommissionMemberFilter(self.request.GET, queryset=queryset)
        return self.filterset.qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter'] = self.filterset
        context['districts'] = DISTRICT_CHOICES  # Add DISTRICT_CHOICES to context
        return context

import django_filters
from .models import CommissionMember, DISTRICT_CHOICES
from django import forms

class CommissionMemberFilter(django_filters.FilterSet):
    full_name = django_filters.CharFilter(lookup_expr='icontains', label='F.I.Sh.',widget=forms.TextInput(attrs={'class': 'form-control'}))
    district = django_filters.ChoiceFilter(choices=DISTRICT_CHOICES, field_name='district_address', label='Tuman',widget=forms.Select(attrs={'class': 'form-control'}))
    nationality = django_filters.CharFilter(lookup_expr='icontains', label='Millati',widget=forms.TextInput(attrs={'class': 'form-control'}))
    gender = django_filters.ChoiceFilter(choices=[('male', 'Erkak'), ('female', 'Ayol')], label='Jinsi',widget=forms.Select(attrs={'class': 'form-control'}))

    class Meta:
        model = CommissionMember
        fields = ['full_name', 'district', 'nationality', 'gender']
        
@login_required
def commission_member_create(request):
    if request.method == 'POST':
        form = CommissionMemberForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "A'zo muvaffaqiyatli qo'shildi!")
            return redirect('commission_member_list')
    else:
        form = CommissionMemberForm()
    
    return render(request, 'elections/commission_members/form.html', {'form': form})

@login_required
def commission_member_update(request, pk):
    member = get_object_or_404(CommissionMember, pk=pk)
    
    if request.method == 'POST':
        form = CommissionMemberForm(request.POST, request.FILES, instance=member)
        if form.is_valid():
            form.save()
            messages.success(request, "A'zo muvaffaqiyatli yangilandi!")
            return redirect('commission_member_list')
    else:
        messages.error(request, "A'zo ma'lumotlarini yangilashda xatolik yuz berdi!")
        form = CommissionMemberForm(instance=member)
    
    return render(request, 'elections/commission_members/form.html', {'form': form, 'member': member})

@login_required
def commission_member_delete(request, pk):
    member = get_object_or_404(CommissionMember, pk=pk)
    
    if request.method == 'POST':
        member.delete()
        messages.success(request, "A'zo muvaffaqiyatli o'chirildi!")
        return redirect('commission_member_list')
    
    return render(request, 'elections/commission_members/delete.html', {'member': member})

import xlsxwriter
from io import BytesIO
from PIL import Image
import os

from django.conf import settings
from django.core.files.storage import default_storage
from django.http import HttpResponse
from .models import CommissionMember
from django.contrib.auth.decorators import login_required

@login_required
def export_commission_members_excel(request):
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet("USK A'zolari")

    # Sarlavhalar
    headers = [
        'Uchastka raqami', 'A\'zoligi', 'F.I.Sh.', 'Tug\'ilgan sanasi',
        'Tug\'ilgan joyi', 'Millati', 'Ma\'lumoti', 'Mutaxassisligi',
        'Ish joyi va lavozimi', 'Telefon raqami', 'Jinsi', 'Yoshi', 'Rasmi'
    ]

    bold = workbook.add_format({'bold': True})
    for col_num, header in enumerate(headers):
        worksheet.write(0, col_num, header, bold)

    # Har bir ustun eni
    for col in range(len(headers) - 1):
        worksheet.set_column(col, col, 20)

    image_col = len(headers) - 1
    worksheet.set_column(image_col, image_col, 15)

    members = CommissionMember.objects.select_related('district').all()
    row_num = 1

    for member in members:
        data = [
            member.district.district_number,
            member.membership_role,
            member.full_name,
            member.birth_date.strftime('%d.%m.%Y') if member.birth_date else '',
            member.birth_place,
            member.nationality,
            member.education,
            member.specialization,
            member.workplace,
            member.phone_number,
            'Erkak' if member.gender == 'male' else 'Ayol',
            member.age or '',
        ]

        for col_num, value in enumerate(data):
            worksheet.write(row_num, col_num, value)

        worksheet.set_row(row_num, 80)  # Rasmga joy bo'lishi uchun satr balandligi

        # Rasmini tayyorlash
        if member.photo and default_storage.exists(member.photo.name):
            image_path = os.path.join(settings.MEDIA_ROOT, member.photo.name)

            # Rasmni Pillow orqali ochamiz va resize qilamiz
            try:
                with Image.open(image_path) as img:
                    # Rasmni qayta o‘lchamlash (masalan 100x100 px)
                    img = img.convert('RGB')
                    img = img.resize((100, 100))

                    # Temporarily saqlaymiz memoryga
                    temp_thumb = BytesIO()
                    img.save(temp_thumb, format='PNG')
                    temp_thumb.seek(0)

                    # Excelga joylaymiz
                    worksheet.insert_image(
                        row_num,
                        image_col,
                        member.full_name + ".png",  # unique image name
                        {
                            'image_data': temp_thumb,
                            'x_offset': 5,
                            'y_offset': 5,
                        }
                    )
            except Exception as e:
                print(f"Rasmni ochishda xatolik: {e}")

        row_num += 1

    workbook.close()
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=commission_members.xlsx'
    return response



from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Spacer
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from io import BytesIO
from PIL import Image as PILImage
import os

from django.conf import settings
from django.core.files.storage import default_storage
from .models import CommissionMember

@login_required
def export_commission_members_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="commission_members.pdf"'

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    members = CommissionMember.objects.select_related('district').all()

    for member in members:
        # Rasm tayyorlash
        img_data = None
        if member.photo and default_storage.exists(member.photo.name):
            image_path = os.path.join(settings.MEDIA_ROOT, member.photo.name)
            try:
                with PILImage.open(image_path) as img:
                    img = img.convert('RGB')
                    img = img.resize((80, 80))  # bir xil o‘lcham
                    img_buffer = BytesIO()
                    img.save(img_buffer, format='PNG')
                    img_buffer.seek(0)
                    img_data = Image(img_buffer, width=80, height=80)
            except Exception as e:
                print(f"Rasmda xatolik: {e}")

        # Matnli ma'lumotlar
        gender = 'Erkak' if member.gender == 'male' else 'Ayol'
        text_data = [
            ['Uchastka raqami', member.district.district_number],
            ['A\'zoligi', member.membership_role],
            ['F.I.Sh.', member.full_name],
            ['Millati', member.nationality],
            ['Ma\'lumoti', member.education],
            ['Yoshi', str(member.age)],
            ['Jinsi', gender],
        ]

        table = Table(text_data, colWidths=[120, 300])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))

        row = [img_data, table] if img_data else [table]
        elements.append(Table([row], colWidths=[100, 400]))
        elements.append(Spacer(1, 15))  # Har a'zo o'rtasida bo‘sh joy

    doc.build(elements)
    buffer.seek(0)
    return HttpResponse(buffer.read(), content_type='application/pdf')


# Election Districts Views
class ElectionDistrictListView(LoginRequiredMixin,ListView):
    model = ElectionDistrict
    template_name = 'elections/districts/list.html'
    context_object_name = 'districts'
    paginate_by = 10
    
    def get_queryset(self):
        queryset = super().get_queryset()
        self.filterset = ElectionDistrictFilter(self.request.GET, queryset=queryset)
        return self.filterset.qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter'] = self.filterset
        return context

@login_required
def election_district_create(request):
    if request.method == 'POST':
        form = ElectionDistrictForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Saylov uchastkasi muvaffaqiyatli qo'shildi!")
            return redirect('election_district_list')
    else:
        form = ElectionDistrictForm()
    
    return render(request, 'elections/districts/form.html', {'form': form})


@login_required
def election_district_update(request, pk):
    district = get_object_or_404(ElectionDistrict, pk=pk)
    
    if request.method == 'POST':
        form = ElectionDistrictForm(request.POST, instance=district)
        if form.is_valid():
            form.save()
            messages.success(request, "Saylov uchastkasi muvaffaqiyatli yangilandi!")
            return redirect('election_district_list')
    else:
        form = ElectionDistrictForm(instance=district)
    
    return render(request, 'elections/districts/form.html', {'form': form, 'district': district})

@login_required
def election_district_delete(request, pk):
    district = get_object_or_404(ElectionDistrict, pk=pk)
    
    if request.method == 'POST':
        district.delete()
        messages.success(request, "Saylov uchastkasi muvaffaqiyatli o'chirildi!")
        return redirect('election_district_list')
    
    return render(request, 'elections/districts/delete.html', {'district': district})

@login_required
def export_districts_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="election_districts.xls"'
    
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Saylov uchastkalari')
    
    # Sheet header
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True
    
    columns = ['Shahar (tuman) nomi', 'Saylov uchastkasi raqami', 'Saylov uchastkasi manzili', 'Saylov uchastkasi chegaralari']
    
    for col_num, column_title in enumerate(columns):
        ws.write(row_num, col_num, column_title, font_style)
    
    # Sheet body
    font_style = xlwt.XFStyle()
    
    rows = ElectionDistrict.objects.all().values_list('city_name', 'district_number', 'address', 'boundaries')
    
    for row in rows:
        row_num += 1
        for col_num, cell_value in enumerate(row):
            ws.write(row_num, col_num, str(cell_value), font_style)
    
    wb.save(response)
    return response

# Representatives Views
class RepresentativeListView(LoginRequiredMixin,ListView):
    model = Representative
    template_name = 'elections/representatives/list.html'
    context_object_name = 'representatives'
    paginate_by = 10
    
    def get_queryset(self):
        queryset = super().get_queryset()
        self.filterset = RepresentativeFilter(self.request.GET, queryset=queryset)
        return self.filterset.qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter'] = self.filterset
        return context

@login_required
def representative_create(request):
    if request.method == 'POST':
        form = RepresentativeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Vakolatli vakil muvaffaqiyatli qo'shildi!")
            return redirect('representative_list')
    else:
        form = RepresentativeForm()
    
    return render(request, 'elections/representatives/form.html', {'form': form})

@login_required
def representative_update(request, pk):
    representative = get_object_or_404(Representative, pk=pk)
    
    if request.method == 'POST':
        form = RepresentativeForm(request.POST, instance=representative)
        if form.is_valid():
            form.save()
            messages.success(request, "Vakolatli vakil muvaffaqiyatli yangilandi!")
            return redirect('representative_list')
    else:
        form = RepresentativeForm(instance=representative)
    
    return render(request, 'elections/representatives/form.html', {'form': form, 'representative': representative})

@login_required
def representative_delete(request, pk):
    representative = get_object_or_404(Representative, pk=pk)
    
    if request.method == 'POST':
        representative.delete()
        messages.success(request, "Vakolatli vakil muvaffaqiyatli o'chirildi!")
        return redirect('representative_list')
    
    return render(request, 'elections/representatives/delete.html', {'representative': representative})

# Observers Views
class ObserverListView(LoginRequiredMixin,ListView):
    model = Observer
    template_name = 'elections/observers/list.html'
    context_object_name = 'observers'
    paginate_by = 10
    
    def get_queryset(self):
        queryset = super().get_queryset()
        self.filterset = ObserverFilter(self.request.GET, queryset=queryset)
        return self.filterset.qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter'] = self.filterset
        return context

@login_required
def observer_create(request):
    if request.method == 'POST':
        form = ObserverForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Kuzatuvchi muvaffaqiyatli qo'shildi!")
            return redirect('observer_list')
    else:
        form = ObserverForm()
    
    return render(request, 'elections/observers/form.html', {'form': form})

def export_representatives_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="representatives.xls"'
    
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Vakolatli vakillar')
    
    # Sheet header
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True
    
    columns = ['Partiya', 'Tuman (shahar) kengashi', 'F.I.Sh.', 'Ish joyi va lavozimi']
    
    for col_num, column_title in enumerate(columns):
        ws.write(row_num, col_num, column_title, font_style)
    
    # Sheet body
    font_style = xlwt.XFStyle()
    
    representatives = Representative.objects.all()
    
    for representative in representatives:
        row_num += 1
        row = [
            representative.get_party_name_display(),  # Use get_party_name_display() instead of dict lookup
            representative.city_council,
            representative.full_name,
            representative.workplace
        ]
        
        for col_num, cell_value in enumerate(row):
            ws.write(row_num, col_num, str(cell_value), font_style)
    
    wb.save(response)
    return response

def export_representatives_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="representatives.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    
    # Define table data
    data = [
        ['Partiya', 'Tuman (shahar) kengashi', 'F.I.Sh.', 'Ish joyi va lavozimi']
    ]
    
    representatives = Representative.objects.all()
    for representative in representatives:
        data.append([
            representative.get_party_name_display(),  # Use get_party_name_display() instead of dict lookup
            representative.city_council,
            representative.full_name,
            representative.workplace
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    return response

def export_observers_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="observers.xls"'
    
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Kuzatuvchilar')
    
    # Sheet header
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True
    
    columns = ['Partiya', 'Uchastka raqami', 'F.I.Sh.', 'Ish joyi va lavozimi']
    
    for col_num, column_title in enumerate(columns):
        ws.write(row_num, col_num, column_title, font_style)
    
    # Sheet body
    font_style = xlwt.XFStyle()
    
    observers = Observer.objects.all().select_related('district')
    
    for observer in observers:
        row_num += 1
        row = [
            observer.get_party_name_display(),  # Use get_party_name_display() instead of dict lookup
            observer.district.district_number,
            observer.full_name,
            observer.workplace
        ]
        
        for col_num, cell_value in enumerate(row):
            ws.write(row_num, col_num, str(cell_value), font_style)
    
    wb.save(response)
    return response

def export_observers_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="observers.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    
    # Define table data
    data = [
        ['Partiya', 'Uchastka raqami', 'F.I.Sh.', 'Ish joyi va lavozimi']
    ]
    
    observers = Observer.objects.all().select_related('district')
    for observer in observers:
        data.append([
            observer.get_party_name_display(),  # Use get_party_name_display() instead of dict lookup
            observer.district.district_number,
            observer.full_name,
            observer.workplace
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    return response

@login_required
def observer_update(request, pk):
    observer = get_object_or_404(Observer, pk=pk)
    
    if request.method == 'POST':
        form = ObserverForm(request.POST, instance=observer)
        if form.is_valid():
            form.save()
            messages.success(request, "Kuzatuvchi muvaffaqiyatli yangilandi!")
            return redirect('observer_list')
    else:
        form = ObserverForm(instance=observer)
    
    return render(request, 'elections/observers/form.html', {'form': form, 'observer': observer})

@login_required
def observer_delete(request, pk):
    observer = get_object_or_404(Observer, pk=pk)
    
    if request.method == 'POST':
        observer.delete()
        messages.success(request, "Kuzatuvchi muvaffaqiyatli o'chirildi!")
        return redirect('observer_list')
    
    return render(request, 'elections/observers/delete.html', {'observer': observer})

from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Sum, Count, Q
from django.contrib import messages
from django.utils import timezone
from datetime import date
from .models import District, Member, DistrictStats
from .forms import MemberForm, MemberCheckForm

def calculate_totals():
    """Calculate total statistics across all districts"""
    totals = {
        'district_count': District.objects.count(),
        'registered_voters': DistrictStats.objects.aggregate(Sum('registered_voters'))['registered_voters__sum'] or 0,
        'total_voters': DistrictStats.objects.aggregate(Sum('total_voters'))['total_voters__sum'] or 0,
        'percentage': 0,  # Will calculate below if registered_voters > 0
        'cancelled_ballots': DistrictStats.objects.aggregate(Sum('cancelled_ballots'))['cancelled_ballots__sum'] or 0,
        
        # Activity hours
        'activity_8': DistrictStats.objects.aggregate(Sum('activity_8'))['activity_8__sum'] or 0,
        'activity_10': DistrictStats.objects.aggregate(Sum('activity_10'))['activity_10__sum'] or 0,
        'activity_12': DistrictStats.objects.aggregate(Sum('activity_12'))['activity_12__sum'] or 0,
        'activity_14': DistrictStats.objects.aggregate(Sum('activity_14'))['activity_14__sum'] or 0,
        'activity_16': DistrictStats.objects.aggregate(Sum('activity_16'))['activity_16__sum'] or 0,
        'activity_18': DistrictStats.objects.aggregate(Sum('activity_18'))['activity_18__sum'] or 0,
        'activity_19': DistrictStats.objects.aggregate(Sum('activity_19'))['activity_19__sum'] or 0,
        'activity_20': DistrictStats.objects.aggregate(Sum('activity_20'))['activity_20__sum'] or 0,
        'activity_percentage': 0,  # Will calculate below
        
        # Specialization counts
        'teacher_count': DistrictStats.objects.aggregate(Sum('teacher_count'))['teacher_count__sum'] or 0,
        'doctor_count': DistrictStats.objects.aggregate(Sum('doctor_count'))['doctor_count__sum'] or 0,
        'engineer_count': DistrictStats.objects.aggregate(Sum('engineer_count'))['engineer_count__sum'] or 0,
        'economist_count': DistrictStats.objects.aggregate(Sum('economist_count'))['economist_count__sum'] or 0,
        'lawyer_count': DistrictStats.objects.aggregate(Sum('lawyer_count'))['lawyer_count__sum'] or 0,
        'agriculture_count': DistrictStats.objects.aggregate(Sum('agriculture_count'))['agriculture_count__sum'] or 0,
        'other_specialization_count': DistrictStats.objects.aggregate(Sum('other_specialization_count'))['other_specialization_count__sum'] or 0,
        'total_specialization_count': DistrictStats.objects.aggregate(Sum('total_specialization_count'))['total_specialization_count__sum'] or 0,
        
        # Nationality counts
        'uzbek_count': DistrictStats.objects.aggregate(Sum('uzbek_count'))['uzbek_count__sum'] or 0,
        'other_nationality_count': DistrictStats.objects.aggregate(Sum('other_nationality_count'))['other_nationality_count__sum'] or 0,
        
        # Education counts
        'higher_education_count': DistrictStats.objects.aggregate(Sum('higher_education_count'))['higher_education_count__sum'] or 0,
        'secondary_education_count': DistrictStats.objects.aggregate(Sum('secondary_education_count'))['secondary_education_count__sum'] or 0,
        
        # Age group counts
        'age_21_30_count': DistrictStats.objects.aggregate(Sum('age_21_30_count'))['age_21_30_count__sum'] or 0,
        'age_31_40_count': DistrictStats.objects.aggregate(Sum('age_31_40_count'))['age_31_40_count__sum'] or 0,
        'age_41_50_count': DistrictStats.objects.aggregate(Sum('age_41_50_count'))['age_41_50_count__sum'] or 0,
        'age_51_60_count': DistrictStats.objects.aggregate(Sum('age_51_60_count'))['age_51_60_count__sum'] or 0,
        'age_above_60_count': DistrictStats.objects.aggregate(Sum('age_above_60_count'))['age_above_60_count__sum'] or 0,
        
        # Gender counts
        'male_count': DistrictStats.objects.aggregate(Sum('male_count'))['male_count__sum'] or 0,
        'female_count': DistrictStats.objects.aggregate(Sum('female_count'))['female_count__sum'] or 0,
        
        # Women in positions
        'female_chairman_count': DistrictStats.objects.aggregate(Sum('female_chairman_count'))['female_chairman_count__sum'] or 0,
        'female_deputy_count': DistrictStats.objects.aggregate(Sum('female_deputy_count'))['female_deputy_count__sum'] or 0,
        'female_secretary_count': DistrictStats.objects.aggregate(Sum('female_secretary_count'))['female_secretary_count__sum'] or 0,
        
        # Other counts
        'disabled_count': DistrictStats.objects.aggregate(Sum('disabled_count'))['disabled_count__sum'] or 0,
        'it_specialist_count': DistrictStats.objects.aggregate(Sum('it_specialist_count'))['it_specialist_count__sum'] or 0,
        'retired_count': DistrictStats.objects.aggregate(Sum('retired_count'))['retired_count__sum'] or 0,
        
        # Language proficiency
        'russian_speakers_count': DistrictStats.objects.aggregate(Sum('russian_speakers_count'))['russian_speakers_count__sum'] or 0,
        'english_speakers_count': DistrictStats.objects.aggregate(Sum('english_speakers_count'))['english_speakers_count__sum'] or 0,
    }
    
    # Calculate percentages if we have registered voters
    if totals['registered_voters'] > 0:
        totals['percentage'] = round((totals['total_voters'] / totals['registered_voters']) * 100, 2)
    
    # Calculate activity percentage
    total_activity = (
        totals['activity_8'] + totals['activity_10'] + totals['activity_12'] + 
        totals['activity_14'] + totals['activity_16'] + totals['activity_18'] + 
        totals['activity_19'] + totals['activity_20']
    )
    if total_activity > 0:
        totals['activity_percentage'] = round((totals['total_voters'] / total_activity) * 100, 2)
    
    return totals

def update_district_stats(district):
    """Update statistics for a specific district based on its members"""
    stats, created = DistrictStats.objects.get_or_create(district=district)
    
    # Get all members for this district
    members = Member.objects.filter(district=district)
    
    # Count members by specialization
    stats.teacher_count = members.filter(specialization='teacher').count()
    stats.doctor_count = members.filter(specialization='doctor').count()
    stats.engineer_count = members.filter(specialization='engineer').count()
    stats.economist_count = members.filter(specialization='economist').count()
    stats.lawyer_count = members.filter(specialization='lawyer').count()
    stats.agriculture_count = members.filter(specialization='agriculture').count()
    stats.other_specialization_count = members.filter(specialization='other').count()
    stats.total_specialization_count = members.count()
    
    # Count members by nationality
    stats.uzbek_count = members.filter(nationality='uzbek').count()
    stats.other_nationality_count = members.exclude(nationality='uzbek').count()
    
    # Count members by education
    stats.higher_education_count = members.filter(education='higher').count()
    stats.secondary_education_count = members.filter(education__in=['secondary_special', 'secondary']).count()
    
    # Count members by age group
    today = date.today()
    stats.age_21_30_count = members.filter(birth_date__gte=date(today.year-30, today.month, today.day), 
                                          birth_date__lte=date(today.year-21, today.month, today.day)).count()
    stats.age_31_40_count = members.filter(birth_date__gte=date(today.year-40, today.month, today.day), 
                                          birth_date__lt=date(today.year-30, today.month, today.day)).count()
    stats.age_41_50_count = members.filter(birth_date__gte=date(today.year-50, today.month, today.day), 
                                          birth_date__lt=date(today.year-40, today.month, today.day)).count()
    stats.age_51_60_count = members.filter(birth_date__gte=date(today.year-60, today.month, today.day), 
                                          birth_date__lt=date(today.year-50, today.month, today.day)).count()
    stats.age_above_60_count = members.filter(birth_date__lt=date(today.year-60, today.month, today.day)).count()
    
    # Count members by gender
    stats.male_count = members.filter(gender='male').count()
    stats.female_count = members.filter(gender='female').count()
    
    # Count women in positions
    stats.female_chairman_count = members.filter(gender='female', position='chairman').count()
    stats.female_deputy_count = members.filter(gender='female', position='deputy').count()
    stats.female_secretary_count = members.filter(gender='female', position='secretary').count()
    
    # Count other categories
    stats.disabled_count = members.filter(is_disabled=True).count()
    stats.it_specialist_count = members.filter(is_it_specialist=True).count()
    stats.retired_count = members.filter(is_retired=True).count()
    
    # Count language proficiency
    stats.russian_speakers_count = members.filter(knows_russian=True).count()
    stats.english_speakers_count = members.filter(knows_english=True).count()
    
    # Update registered voters count (this might come from another source in a real app)
    stats.registered_voters = members.count()
    stats.total_voters = members.filter(participation_status='participated').count()
    
    # Calculate percentage if we have registered voters
    if stats.registered_voters > 0:
        stats.percentage = round((stats.total_voters / stats.registered_voters) * 100, 2)
    
    stats.save()
    return stats


@login_required
def home(request):
    districts = District.objects.all()
    all_members = Member.objects.all()
    
    # Calculate totals for all statistics
    context = {
        'districts': districts,
        'all_members': all_members,
        'total_precincts':ElectionDistrict.objects.count(),  # Total precincts across all districts
        'participated_count': all_members.filter(participation_status='participated').count(),
        'not_participated_count': all_members.filter(participation_status='not_participated').count(),
        'public_education_count': all_members.filter(activity_field='public_education').count(),
        'higher_education_count': all_members.filter(activity_field='higher_education').count(),
        'government_count': all_members.filter(activity_field='government').count(),
        'self_governance_count': all_members.filter(activity_field='self_governance').count(),
        'ngo_count': all_members.filter(activity_field='ngo').count(),
        'private_sector_count': all_members.filter(activity_field='private_sector').count(),
        'unemployed_count': all_members.filter(activity_field='unemployed').count(),
        'teacher_count': all_members.filter(specialization='teacher').count(),
        'doctor_count': all_members.filter(specialization='doctor').count(),
        'engineer_count': all_members.filter(specialization='engineer').count(),
        'economist_count': all_members.filter(specialization='economist').count(),
        'journalist_count': all_members.filter(specialization='journalist').count(),
        'lawyer_count': all_members.filter(specialization='lawyer').count(),
        'agriculture_count': all_members.filter(specialization='agriculture').count(),
        'other_spec_count': all_members.filter(specialization='other').count(),
        'uzbek_count': all_members.filter(nationality='uzbek').count(),
        'karakalpak_count': all_members.filter(nationality='karakalpak').count(),
        'kazakh_count': all_members.filter(nationality='kazakh').count(),
        'russian_count': all_members.filter(nationality='russian').count(),
        'tajik_count': all_members.filter(nationality='tajik').count(),
        'korean_count': all_members.filter(nationality='korean').count(),
        'other_nat_count': all_members.filter(nationality='other').count(),
        'higher_count': all_members.filter(education='higher').count(),
        'secondary_special_count': all_members.filter(education='secondary_special').count(),
        'secondary_count': all_members.filter(education='secondary').count(),
        'age_21_30_count': sum(1 for member in all_members if 21 <= (datetime.date.today().year - member.birth_date.year) <= 30),
        'age_31_40_count': sum(1 for member in all_members if 31 <= (datetime.date.today().year - member.birth_date.year) <= 40),
        'age_41_50_count': sum(1 for member in all_members if 41 <= (datetime.date.today().year - member.birth_date.year) <= 50),
        'age_51_60_count': sum(1 for member in all_members if 51 <= (datetime.date.today().year - member.birth_date.year) <= 60),
        'age_60_plus_count': sum(1 for member in all_members if (datetime.date.today().year - member.birth_date.year) > 60),
        'female_chairman_count': all_members.filter(gender='female', position='chairman').count(),
        'female_deputy_count': all_members.filter(gender='female', position='deputy').count(),
        'female_secretary_count': all_members.filter(gender='female', position='secretary').count(),
        'disabled_count': all_members.filter(is_disabled=True).count(),
        'it_specialist_count': all_members.filter(is_it_specialist=True).count(),
        'retired_count': all_members.filter(is_retired=True).count(),
        'russian_speakers_count': all_members.filter(knows_russian=True).count(),
        'english_speakers_count': all_members.filter(knows_english=True).count(),
    }
    
    # Calculate statistics for each district
    for district in districts:
        members = Member.objects.filter(district=district)
        district.precinct_count = district.precincts.count()  # Count precincts for this district
        district.voter_count = members.count()  # Count registered voters
        
        # Status counts
        district.participated_count = members.filter(participation_status='participated').count()
        district.not_participated_count = members.filter(participation_status='not_participated').count()
        
        # Activity fields
        district.public_education_count = members.filter(activity_field='public_education').count()
        district.higher_education_count = members.filter(activity_field='higher_education').count()
        district.government_count = members.filter(activity_field='government').count()
        district.self_governance_count = members.filter(activity_field='self_governance').count()
        district.ngo_count = members.filter(activity_field='ngo').count()
        district.private_sector_count = members.filter(activity_field='private_sector').count()
        district.unemployed_count = members.filter(activity_field='unemployed').count()
        
        # Specialization
        district.teacher_count = members.filter(specialization='teacher').count()
        district.doctor_count = members.filter(specialization='doctor').count()
        district.engineer_count = members.filter(specialization='engineer').count()
        district.economist_count = members.filter(specialization='economist').count()
        district.journalist_count = members.filter(specialization='journalist').count()
        district.lawyer_count = members.filter(specialization='lawyer').count()
        district.agriculture_count = members.filter(specialization='agriculture').count()
        district.other_spec_count = members.filter(specialization='other').count()
        
        # Nationality
        district.uzbek_count = members.filter(nationality='uzbek').count()
        district.karakalpak_count = members.filter(nationality='karakalpak').count()
        district.kazakh_count = members.filter(nationality='kazakh').count()
        district.russian_count = members.filter(nationality='russian').count()
        district.tajik_count = members.filter(nationality='tajik').count()
        district.korean_count = members.filter(nationality='korean').count()
        district.other_nat_count = members.filter(nationality='other').count()
        
        # Education
        district.higher_count = members.filter(education='higher').count()
        district.secondary_special_count = members.filter(education='secondary_special').count()
        district.secondary_count = members.filter(education='secondary').count()
        
        # Age groups
        district.age_21_30_count = sum(1 for member in members if 21 <= (datetime.date.today().year - member.birth_date.year) <= 30)
        district.age_31_40_count = sum(1 for member in members if 31 <= (datetime.date.today().year - member.birth_date.year) <= 40)
        district.age_41_50_count = sum(1 for member in members if 41 <= (datetime.date.today().year - member.birth_date.year) <= 50)
        district.age_51_60_count = sum(1 for member in members if 51 <= (datetime.date.today().year - member.birth_date.year) <= 60)
        district.age_60_plus_count = sum(1 for member in members if (datetime.date.today().year - member.birth_date.year) > 60)
        
        # Women in positions
        district.female_chairman_count = members.filter(gender='female', position='chairman').count()
        district.female_deputy_count = members.filter(gender='female', position='deputy').count()
        district.female_secretary_count = members.filter(gender='female', position='secretary').count()
        
        # Other statistics
        district.disabled_count = members.filter(is_disabled=True).count()
        district.it_specialist_count = members.filter(is_it_specialist=True).count()
        district.retired_count = members.filter(is_retired=True).count()
        
        # Language proficiency
        district.russian_speakers_count = members.filter(knows_russian=True).count()
        district.english_speakers_count = members.filter(knows_english=True).count()
    
    return render(request, 'elections/home.html', context)






def add_member(request):
    """View for adding a new member"""
    if request.method == 'POST':
        form = MemberForm(request.POST)
        if form.is_valid():
            member = form.save()
            
            # Update district statistics
            update_district_stats(member.district)
            
            messages.success(request, f"A'zo {member.first_name} {member.last_name} muvaffaqiyatli qo'shildi!")
            return redirect('home')
    else:
        form = MemberForm()
    
    return render(request, 'elections/add.html', {'form': form})

def check_member(request):
    """View for checking if a member exists"""
    members = None
    
    if request.method == 'POST':
        form = MemberCheckForm(request.POST)
        if form.is_valid():
            first_name = form.cleaned_data['first_name']
            last_name = form.cleaned_data['last_name']
            district = form.cleaned_data['district']
            
            # Build query
            query = Q(first_name__icontains=first_name) & Q(last_name__icontains=last_name)
            if district:
                query &= Q(district=district)
            
            members = Member.objects.filter(query)
            
            if not members:
                messages.info(request, "Bunday a'zo topilmadi.")
    else:
        form = MemberCheckForm()
    
    return render(request, 'elections/check.html', {
        'form': form,
        'members': members
    })

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import datetime
from .models import District, Member  # Update this to match your actual model names

def export_excel(request):
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Saylov Ma'lumotlari"
    
    # Define styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    subheader_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    total_row_fill = PatternFill(start_color="E6F7FF", end_color="E6F7FF", fill_type="solid")
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Add main header row with merged cells for column groups
    # Row 1: Main headers with merged cells
    row_idx = 1
    
    # Basic columns (no merging needed)
    ws.cell(row=row_idx, column=1, value="№")
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx+1, end_column=1)
    
    ws.cell(row=row_idx, column=2, value="Hududlar")
    ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx+1, end_column=2)
    
    ws.cell(row=row_idx, column=3, value="Uchastkalar soni")
    ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx+1, end_column=3)
    
    ws.cell(row=row_idx, column=4, value="Ro'yxatga olingan saylovchilar soni")
    ws.merge_cells(start_row=row_idx, start_column=4, end_row=row_idx+1, end_column=4)
    
    # Holat (Status) - 2 columns
    ws.cell(row=row_idx, column=5, value="Holat")
    ws.merge_cells(start_row=row_idx, start_column=5, end_row=row_idx, end_column=6)
    
    # Faoliyat sohasi (Activity field) - 7 columns
    ws.cell(row=row_idx, column=7, value="Faoliyat sohasi")
    ws.merge_cells(start_row=row_idx, start_column=7, end_row=row_idx, end_column=13)
    
    # Mutaxassisligi (Specialization) - 8 columns
    ws.cell(row=row_idx, column=14, value="Mutaxassisligi")
    ws.merge_cells(start_row=row_idx, start_column=14, end_row=row_idx, end_column=21)
    
    # Millati (Nationality) - 7 columns
    ws.cell(row=row_idx, column=22, value="Millati")
    ws.merge_cells(start_row=row_idx, start_column=22, end_row=row_idx, end_column=28)
    
    # Ma'lumoti (Education) - 3 columns
    ws.cell(row=row_idx, column=29, value="Ma'lumoti")
    ws.merge_cells(start_row=row_idx, start_column=29, end_row=row_idx, end_column=31)
    
    # Yoshi (Age) - 5 columns
    ws.cell(row=row_idx, column=32, value="Yoshi")
    ws.merge_cells(start_row=row_idx, start_column=32, end_row=row_idx, end_column=36)
    
    # Shu jumlada (Including) - 3 columns
    ws.cell(row=row_idx, column=37, value="Shu jumlada")
    ws.merge_cells(start_row=row_idx, start_column=37, end_row=row_idx, end_column=39)
    
    # Nogironligi tegishli shaxslar (Disabled persons) - 1 column
    ws.cell(row=row_idx, column=40, value="Nogironligi tegishli shaxslar")
    ws.merge_cells(start_row=row_idx, start_column=40, end_row=row_idx+1, end_column=40)
    
    # Axborot-kommunikatsiya texnologiyalari yaratish (IT specialists) - 1 column
    ws.cell(row=row_idx, column=41, value="Axborot-kommunikatsiya texnologiyalari yaratish")
    ws.merge_cells(start_row=row_idx, start_column=41, end_row=row_idx+1, end_column=41)
    
    # Nafaqadagilar (Retired) - 1 column
    ws.cell(row=row_idx, column=42, value="Nafaqadagilar")
    ws.merge_cells(start_row=row_idx, start_column=42, end_row=row_idx+1, end_column=42)
    
    # Xorijiy tillar (Foreign languages) - 2 columns
    ws.cell(row=row_idx, column=43, value="Xorijiy tillar")
    ws.merge_cells(start_row=row_idx, start_column=43, end_row=row_idx, end_column=44)
    
    # Apply styles to main header row
    for col_idx in range(1, 45):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Row 2: Subheaders
    row_idx = 2
    
    # Holat (Status) subheaders
    ws.cell(row=row_idx, column=5, value="Avval qatnashgan")
    ws.cell(row=row_idx, column=6, value="Avval qatnashmagan")
    
    # Faoliyat sohasi (Activity field) subheaders
    ws.cell(row=row_idx, column=7, value="Xalq ta'limi tizimi hodimi")
    ws.cell(row=row_idx, column=8, value="Oliy ta'lim tizimi hodimi")
    ws.cell(row=row_idx, column=9, value="Boshqa davlat boshqaruvlari va muassasalar xodimlari")
    ws.cell(row=row_idx, column=10, value="O'zini o'zi boshqarish organi vakillari")
    ws.cell(row=row_idx, column=11, value="Nodavlat notijorat tashkilotlar xodimlari")
    ws.cell(row=row_idx, column=12, value="Boshqa nodavlat mavjudligi sohasi vakillari")
    ws.cell(row=row_idx, column=13, value="Vaqtcha ishsizlar")
    
    # Mutaxassisligi (Specialization) subheaders
    ws.cell(row=row_idx, column=14, value="O'qituvchi")
    ws.cell(row=row_idx, column=15, value="Shifokor")
    ws.cell(row=row_idx, column=16, value="Muhandis")
    ws.cell(row=row_idx, column=17, value="Iqtisodchi")
    ws.cell(row=row_idx, column=18, value="Jurnalist")
    ws.cell(row=row_idx, column=19, value="Huquqshunos")
    ws.cell(row=row_idx, column=20, value="Qishloq xo'jaligi")
    ws.cell(row=row_idx, column=21, value="Boshqa")
    
    # Millati (Nationality) subheaders
    ws.cell(row=row_idx, column=22, value="O'zbek")
    ws.cell(row=row_idx, column=23, value="Qoraqalpoq")
    ws.cell(row=row_idx, column=24, value="Qozoq")
    ws.cell(row=row_idx, column=25, value="Rus")
    ws.cell(row=row_idx, column=26, value="Tojik")
    ws.cell(row=row_idx, column=27, value="Koreys")
    ws.cell(row=row_idx, column=28, value="Boshqa")
    
    # Ma'lumoti (Education) subheaders
    ws.cell(row=row_idx, column=29, value="Oliy")
    ws.cell(row=row_idx, column=30, value="O'rta maxsus")
    ws.cell(row=row_idx, column=31, value="O'rta")
    
    # Yoshi (Age) subheaders
    ws.cell(row=row_idx, column=32, value="21-30 gacha")
    ws.cell(row=row_idx, column=33, value="31-40 gacha")
    ws.cell(row=row_idx, column=34, value="41-50 gacha")
    ws.cell(row=row_idx, column=35, value="51-60 gacha")
    ws.cell(row=row_idx, column=36, value="60 yoshdan kattalar")
    
    # Shu jumlada (Including) subheaders
    ws.cell(row=row_idx, column=37, value="Rais ayollar")
    ws.cell(row=row_idx, column=38, value="Rais o'rinbosari ayollar")
    ws.cell(row=row_idx, column=39, value="Kotib ayollar")
    
    # Xorijiy tillar (Foreign languages) subheaders
    ws.cell(row=row_idx, column=43, value="Mukammal rus tilini biladi")
    ws.cell(row=row_idx, column=44, value="Mukammal ingliz tilini biladi")
    
    # Apply styles to subheader row
    for col_idx in range(1, 45):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = header_font
        cell.fill = subheader_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Get data from the database
    districts = District.objects.all()
    
    # Add total row first
    all_members = Member.objects.all()
    totals = {
        'district_count': districts.count(),
        'registered_members': all_members.count(),
        'participated': all_members.filter(participation_status='participated').count(),
        'not_participated': all_members.filter(participation_status='not_participated').count(),
        
        # Activity fields
        'public_education': all_members.filter(activity_field='public_education').count(),
        'higher_education': all_members.filter(activity_field='higher_education').count(),
        'government': all_members.filter(activity_field='government').count(),
        'self_governance': all_members.filter(activity_field='self_governance').count(),
        'ngo': all_members.filter(activity_field='ngo').count(),
        'private_sector': all_members.filter(activity_field='private_sector').count(),
        'unemployed': all_members.filter(activity_field='unemployed').count(),
        
        # Specialization
        'teacher': all_members.filter(specialization='teacher').count(),
        'doctor': all_members.filter(specialization='doctor').count(),
        'engineer': all_members.filter(specialization='engineer').count(),
        'economist': all_members.filter(specialization='economist').count(),
        'journalist': all_members.filter(specialization='journalist').count(),
        'lawyer': all_members.filter(specialization='lawyer').count(),
        'agriculture': all_members.filter(specialization='agriculture').count(),
        'other_spec': all_members.filter(specialization='other').count(),
        
        # Nationality
        'uzbek': all_members.filter(nationality='uzbek').count(),
        'karakalpak': all_members.filter(nationality='karakalpak').count(),
        'kazakh': all_members.filter(nationality='kazakh').count(),
        'russian': all_members.filter(nationality='russian').count(),
        'tajik': all_members.filter(nationality='tajik').count(),
        'korean': all_members.filter(nationality='korean').count(),
        'other_nat': all_members.filter(nationality='other').count(),
        
        # Education
        'higher': all_members.filter(education='higher').count(),
        'secondary_special': all_members.filter(education='secondary_special').count(),
        'secondary': all_members.filter(education='secondary').count(),
        
        # Age groups - assuming you have an age_group method in your Member model
        'age_21_30': sum(1 for member in all_members if 21 <= (datetime.date.today().year - member.birth_date.year) <= 30),
        'age_31_40': sum(1 for member in all_members if 31 <= (datetime.date.today().year - member.birth_date.year) <= 40),
        'age_41_50': sum(1 for member in all_members if 41 <= (datetime.date.today().year - member.birth_date.year) <= 50),
        'age_51_60': sum(1 for member in all_members if 51 <= (datetime.date.today().year - member.birth_date.year) <= 60),
        'age_60_plus': sum(1 for member in all_members if (datetime.date.today().year - member.birth_date.year) > 60),
        
        # Women in positions
        'female_chairman': all_members.filter(gender='female', position='chairman').count(),
        'female_deputy': all_members.filter(gender='female', position='deputy').count(),
        'female_secretary': all_members.filter(gender='female', position='secretary').count(),
        
        # Other statistics
        'disabled': all_members.filter(is_disabled=True).count(),
        'it_specialist': all_members.filter(is_it_specialist=True).count(),
        'retired': all_members.filter(is_retired=True).count(),
        
        # Language proficiency
        'russian_speakers': all_members.filter(knows_russian=True).count(),
        'english_speakers': all_members.filter(knows_english=True).count(),
    }
    
    # Add total row
    row_idx = 3
    ws.cell(row=row_idx, column=1, value="")
    ws.cell(row=row_idx, column=2, value="Jami").font = Font(bold=True, color="0000FF")
    
    # Fill in total row data
    ws.cell(row=row_idx, column=3, value=totals['district_count'])
    ws.cell(row=row_idx, column=4, value=totals['registered_members'])
    ws.cell(row=row_idx, column=5, value=totals['participated'])
    ws.cell(row=row_idx, column=6, value=totals['not_participated'])
    ws.cell(row=row_idx, column=7, value=totals['public_education'])
    ws.cell(row=row_idx, column=8, value=totals['higher_education'])
    ws.cell(row=row_idx, column=9, value=totals['government'])
    ws.cell(row=row_idx, column=10, value=totals['self_governance'])
    ws.cell(row=row_idx, column=11, value=totals['ngo'])
    ws.cell(row=row_idx, column=12, value=totals['private_sector'])
    ws.cell(row=row_idx, column=13, value=totals['unemployed'])
    ws.cell(row=row_idx, column=14, value=totals['teacher'])
    ws.cell(row=row_idx, column=15, value=totals['doctor'])
    ws.cell(row=row_idx, column=16, value=totals['engineer'])
    ws.cell(row=row_idx, column=17, value=totals['economist'])
    ws.cell(row=row_idx, column=18, value=totals['journalist'])
    ws.cell(row=row_idx, column=19, value=totals['lawyer'])
    ws.cell(row=row_idx, column=20, value=totals['agriculture'])
    ws.cell(row=row_idx, column=21, value=totals['other_spec'])
    ws.cell(row=row_idx, column=22, value=totals['uzbek'])
    ws.cell(row=row_idx, column=23, value=totals['karakalpak'])
    ws.cell(row=row_idx, column=24, value=totals['kazakh'])
    ws.cell(row=row_idx, column=25, value=totals['russian'])
    ws.cell(row=row_idx, column=26, value=totals['tajik'])
    ws.cell(row=row_idx, column=27, value=totals['korean'])
    ws.cell(row=row_idx, column=28, value=totals['other_nat'])
    ws.cell(row=row_idx, column=29, value=totals['higher'])
    ws.cell(row=row_idx, column=30, value=totals['secondary_special'])
    ws.cell(row=row_idx, column=31, value=totals['secondary'])
    ws.cell(row=row_idx, column=32, value=totals['age_21_30'])
    ws.cell(row=row_idx, column=33, value=totals['age_31_40'])
    ws.cell(row=row_idx, column=34, value=totals['age_41_50'])
    ws.cell(row=row_idx, column=35, value=totals['age_51_60'])
    ws.cell(row=row_idx, column=36, value=totals['age_60_plus'])
    ws.cell(row=row_idx, column=37, value=totals['female_chairman'])
    ws.cell(row=row_idx, column=38, value=totals['female_deputy'])
    ws.cell(row=row_idx, column=39, value=totals['female_secretary'])
    ws.cell(row=row_idx, column=40, value=totals['disabled'])
    ws.cell(row=row_idx, column=41, value=totals['it_specialist'])
    ws.cell(row=row_idx, column=42, value=totals['retired'])
    ws.cell(row=row_idx, column=43, value=totals['russian_speakers'])
    ws.cell(row=row_idx, column=44, value=totals['english_speakers'])
    
    # Apply styles to total row
    for col_idx in range(1, 45):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = total_row_fill
        cell.alignment = center_alignment
        cell.border = border
        cell.font = Font(bold=True)
    
    # Add district rows
    row_idx = 4
    for idx, district in enumerate(districts, 1):
        members = Member.objects.filter(district=district)
        
        ws.cell(row=row_idx, column=1, value=idx)
        ws.cell(row=row_idx, column=2, value=district.name)
        ws.cell(row=row_idx, column=3, value=1)  # Each district counts as 1
        ws.cell(row=row_idx, column=4, value=members.count())
        ws.cell(row=row_idx, column=5, value=members.filter(participation_status='participated').count())
        ws.cell(row=row_idx, column=6, value=members.filter(participation_status='not_participated').count())
        ws.cell(row=row_idx, column=7, value=members.filter(activity_field='public_education').count())
        ws.cell(row=row_idx, column=8, value=members.filter(activity_field='higher_education').count())
        ws.cell(row=row_idx, column=9, value=members.filter(activity_field='government').count())
        ws.cell(row=row_idx, column=10, value=members.filter(activity_field='self_governance').count())
        ws.cell(row=row_idx, column=11, value=members.filter(activity_field='ngo').count())
        ws.cell(row=row_idx, column=12, value=members.filter(activity_field='private_sector').count())
        ws.cell(row=row_idx, column=13, value=members.filter(activity_field='unemployed').count())
        ws.cell(row=row_idx, column=14, value=members.filter(specialization='teacher').count())
        ws.cell(row=row_idx, column=15, value=members.filter(specialization='doctor').count())
        ws.cell(row=row_idx, column=16, value=members.filter(specialization='engineer').count())
        ws.cell(row=row_idx, column=17, value=members.filter(specialization='economist').count())
        ws.cell(row=row_idx, column=18, value=members.filter(specialization='journalist').count())
        ws.cell(row=row_idx, column=19, value=members.filter(specialization='lawyer').count())
        ws.cell(row=row_idx, column=20, value=members.filter(specialization='agriculture').count())
        ws.cell(row=row_idx, column=21, value=members.filter(specialization='other').count())
        ws.cell(row=row_idx, column=22, value=members.filter(nationality='uzbek').count())
        ws.cell(row=row_idx, column=23, value=members.filter(nationality='karakalpak').count())
        ws.cell(row=row_idx, column=24, value=members.filter(nationality='kazakh').count())
        ws.cell(row=row_idx, column=25, value=members.filter(nationality='russian').count())
        ws.cell(row=row_idx, column=26, value=members.filter(nationality='tajik').count())
        ws.cell(row=row_idx, column=27, value=members.filter(nationality='korean').count())
        ws.cell(row=row_idx, column=28, value=members.filter(nationality='other').count())
        ws.cell(row=row_idx, column=29, value=members.filter(education='higher').count())
        ws.cell(row=row_idx, column=30, value=members.filter(education='secondary_special').count())
        ws.cell(row=row_idx, column=31, value=members.filter(education='secondary').count())
        
        # Age groups
        ws.cell(row=row_idx, column=32, value=sum(1 for member in members if 21 <= (datetime.date.today().year - member.birth_date.year) <= 30))
        ws.cell(row=row_idx, column=33, value=sum(1 for member in members if 31 <= (datetime.date.today().year - member.birth_date.year) <= 40))
        ws.cell(row=row_idx, column=34, value=sum(1 for member in members if 41 <= (datetime.date.today().year - member.birth_date.year) <= 50))
        ws.cell(row=row_idx, column=35, value=sum(1 for member in members if 51 <= (datetime.date.today().year - member.birth_date.year) <= 60))
        ws.cell(row=row_idx, column=36, value=sum(1 for member in members if (datetime.date.today().year - member.birth_date.year) > 60))
        
        ws.cell(row=row_idx, column=37, value=members.filter(gender='female', position='chairman').count())
        ws.cell(row=row_idx, column=38, value=members.filter(gender='female', position='deputy').count())
        ws.cell(row=row_idx, column=39, value=members.filter(gender='female', position='secretary').count())
        ws.cell(row=row_idx, column=40, value=members.filter(is_disabled=True).count())
        ws.cell(row=row_idx, column=41, value=members.filter(is_it_specialist=True).count())
        ws.cell(row=row_idx, column=42, value=members.filter(is_retired=True).count())
        ws.cell(row=row_idx, column=43, value=members.filter(knows_russian=True).count())
        ws.cell(row=row_idx, column=44, value=members.filter(knows_english=True).count())
        
        # Apply styles to district row
        for col_idx in range(1, 45):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = center_alignment
            cell.border = border
            
            # Apply alternating row colors
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        
        row_idx += 1
    
    # Set column widths
    for col_idx in range(1, 45):
        column_letter = get_column_letter(col_idx)
        if col_idx == 1:  # № column
            ws.column_dimensions[column_letter].width = 5
        elif col_idx == 2:  # Hududlar column
            ws.column_dimensions[column_letter].width = 20
        else:
            ws.column_dimensions[column_letter].width = 12
    
    # Create response with Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=saylov_malumotlari_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    # Save the workbook to the response
    wb.save(response)
    return response


